import io
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment,
    Border, Side
)
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from weasyprint import HTML


# ─────────────────────────────────────────
# EXCEL REPORT GENERATORS
# ─────────────────────────────────────────

def style_header_row(ws, row, cols, fill_color="1A5276"):
    """Apply header styling to a row"""
    fill = PatternFill(
        start_color=fill_color,
        end_color=fill_color,
        fill_type="solid"
    )
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def auto_width_columns(ws):
    """Auto-fit column widths"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


def generate_spend_excel_report(start_date=None, end_date=None):
    """Generate full spend Excel report with multiple sheets"""
    from .queries import (
        get_total_spend_summary,
        get_spend_by_department,
        get_spend_by_month,
        get_vendor_performance_summary,
        get_overdue_invoices_report,
        get_payment_method_breakdown,
    )
    from django.utils import timezone

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Summary"

    summary = get_total_spend_summary(start_date, end_date)
    generated_at = timezone.now().strftime('%d %b %Y %H:%M')

    ws1['A1'] = 'PROCUREMENT SPEND REPORT'
    ws1['A1'].font = Font(size=16, bold=True, color='1A5276')
    ws1['A2'] = f'Generated: {generated_at}'
    ws1['A2'].font = Font(color='888888', italic=True)

    if start_date or end_date:
        period = f"{start_date or 'Beginning'} to {end_date or 'Now'}"
        ws1['A3'] = f'Period: {period}'

    ws1.append([])
    ws1.append(['Metric', 'Value'])
    style_header_row(ws1, ws1.max_row, 2)

    rows = [
        ['Total Spend', f"${summary['total_spend']:,.2f}"],
        ['Total Transactions', summary['total_transactions']],
        ['Average Transaction', f"${summary['average_transaction']:,.2f}"],
        ['Largest Payment', f"${summary['largest_payment']:,.2f}"],
        ['Smallest Payment', f"${summary['smallest_payment']:,.2f}"],
    ]
    for row in rows:
        ws1.append(row)

    auto_width_columns(ws1)

    # ── Sheet 2: Spend by Department ──
    ws2 = wb.create_sheet("By Department")
    ws2.append(['Department', 'Total Spend', 'Transactions', 'Average Spend'])
    style_header_row(ws2, 1, 4)

    dept_data = get_spend_by_department(start_date, end_date)
    for item in dept_data:
        ws2.append([
            item['department_name'],
            float(item['total_spend'] or 0),
            item['transaction_count'],
            float(item['average_spend'] or 0),
        ])
    auto_width_columns(ws2)

    # ── Sheet 3: Monthly Spend ──
    ws3 = wb.create_sheet("Monthly Trend")
    ws3.append(['Month', 'Total Spend', 'Transactions'])
    style_header_row(ws3, 1, 3)

    monthly_data = get_spend_by_month()
    for item in monthly_data:
        ws3.append([
            item['month'].strftime('%B %Y'),
            float(item['total_spend'] or 0),
            item['transaction_count'],
        ])
    auto_width_columns(ws3)

    # ── Sheet 4: Vendor Performance ──
    ws4 = wb.create_sheet("Vendor Performance")
    ws4.append([
        'Company', 'Country', 'Total Bids', 'Won',
        'Win Rate %', 'Total POs', 'Delivered',
        'Delivery Rate %', 'Total Invoiced'
    ])
    style_header_row(ws4, 1, 9)

    vendor_data = get_vendor_performance_summary(start_date, end_date)
    for v in vendor_data:
        ws4.append([
            v['company_name'],
            v['country'],
            v['total_bids'],
            v['awarded_bids'],
            v['win_rate_percent'],
            v['total_purchase_orders'],
            v['delivered_orders'],
            v['delivery_rate_percent'],
            v['total_invoiced_amount'],
        ])
    auto_width_columns(ws4)

    # ── Sheet 5: Overdue Invoices ──
    ws5 = wb.create_sheet("Overdue Invoices")
    ws5.append([
        'Invoice Number', 'Vendor', 'PO Number',
        'Amount', 'Due Date', 'Days Overdue', 'Status'
    ])
    style_header_row(ws5, 1, 7, fill_color='922B21')

    overdue = get_overdue_invoices_report()
    for inv in overdue:
        ws5.append([
            inv['invoice_number'],
            inv['vendor'],
            inv['po_number'],
            float(inv['amount']),
            str(inv['due_date']),
            inv['days_overdue'],
            inv['status'],
        ])
    auto_width_columns(ws5)

    # ── Sheet 6: Payment Methods ──
    ws6 = wb.create_sheet("Payment Methods")
    ws6.append(['Payment Method', 'Count', 'Total Amount'])
    style_header_row(ws6, 1, 3)

    payment_data = get_payment_method_breakdown()
    for item in payment_data:
        ws6.append([
            item['payment_method'],
            item['count'],
            float(item['total_amount'] or 0),
        ])
    auto_width_columns(ws6)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_vendor_performance_excel(start_date=None, end_date=None):
    """Standalone vendor performance Excel report"""
    from .queries import get_vendor_performance_summary

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendor Performance"

    ws['A1'] = 'VENDOR PERFORMANCE REPORT'
    ws['A1'].font = Font(size=16, bold=True, color='1A5276')
    ws.append([])

    headers = [
        'Rank', 'Company', 'City', 'Country', 'Rating',
        'Total Bids', 'Bids Won', 'Win Rate %',
        'Total POs', 'Delivered', 'Delivery Rate %',
        'Total Invoiced ($)'
    ]
    ws.append(headers)
    style_header_row(ws, ws.max_row, len(headers))

    vendors = get_vendor_performance_summary(start_date, end_date)
    for rank, v in enumerate(vendors, start=1):
        ws.append([
            rank,
            v['company_name'],
            v['city'],
            v['country'],
            v['rating'],
            v['total_bids'],
            v['awarded_bids'],
            v['win_rate_percent'],
            v['total_purchase_orders'],
            v['delivered_orders'],
            v['delivery_rate_percent'],
            v['total_invoiced_amount'],
        ])

    auto_width_columns(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────
# PDF REPORT GENERATOR
# ─────────────────────────────────────────

def generate_spend_pdf_report(start_date=None, end_date=None):
    """Generate PDF spend report using WeasyPrint"""
    from .queries import (
        get_total_spend_summary,
        get_spend_by_department,
        get_spend_by_month,
        get_overdue_invoices_report,
    )
    from django.utils import timezone

    summary = get_total_spend_summary(start_date, end_date)
    dept_data = list(get_spend_by_department(start_date, end_date))
    monthly_data = list(get_spend_by_month())
    overdue = get_overdue_invoices_report()

    html_string = render_to_string('pdf/spend_report.html', {
        'summary': summary,
        'dept_data': dept_data,
        'monthly_data': monthly_data,
        'overdue': overdue,
        'generated_at': timezone.now().strftime('%d %b %Y %H:%M'),
        'period': f"{start_date or 'All time'} to {end_date or 'Present'}",
    })

    pdf_bytes = HTML(string=html_string).write_pdf()
    return pdf_bytes
